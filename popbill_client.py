"""
팝빌 홈택스 매입 세금계산서 GUI 다운로더

실행: python popbill_client.py

기능:
  - 조회 기간 입력 (기본: 이번 달)
  - 조회 → 매입 세금계산서 목록 표시
  - 체크박스로 선택 (전체선택/해제 버튼)
  - 선택 건 PDF 저장 (downloads/<승인번호>.pdf)

캐시: 같은 기간 60분 내 재조회는 cache/ 폴더 재사용
"""
from __future__ import annotations

import asyncio
import json
import os
import sys
import threading
import time
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from popbill import HTTaxinvoiceService, PopbillException

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ============================================================
# 설정
# ============================================================
def _load_config() -> dict:
    """config.json 에서 민감 정보 로드 (exe 옆 또는 소스 옆)."""
    candidates = []
    if hasattr(sys, "_MEIPASS"):
        candidates.append(Path(sys.executable).parent / "config.json")
    candidates.append(Path(__file__).parent / "config.json")
    for p in candidates:
        if p.exists():
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    return {}

_CFG = _load_config()
LINK_ID = _CFG.get("LINK_ID", "")
SECRET_KEY = _CFG.get("SECRET_KEY", "")
CORP_NUM = _CFG.get("CORP_NUM", "")
USER_ID = _CFG.get("USER_ID") or None
IS_TEST = _CFG.get("IS_TEST", True)
INVOICE_TYPE = "BUY"  # BUY=매입, SELL=매출

CACHE_TTL_MIN = 60

APP_VERSION = "1.1.0"
GITHUB_REPO = "soluetait/Tax"


def app_data_dir() -> Path:
    """배포용 설정/캐시 저장 위치 (%APPDATA%\\PopbillTaxInvoice)."""
    base = os.environ.get("APPDATA") or str(Path.home() / "AppData" / "Roaming")
    d = Path(base) / "PopbillTaxInvoice"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _init_file_log() -> Path:
    """파일 로깅 초기화 — exe 모드에서 print/traceback 을 파일로 리다이렉트."""
    try:
        log_path = Path(
            os.environ.get("APPDATA") or str(Path.home() / "AppData" / "Roaming")
        ) / "PopbillTaxInvoice" / "app.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        # stdout / stderr 를 파일로 리다이렉트
        f = open(log_path, "a", encoding="utf-8", buffering=1)
        f.write(f"\n\n=== {datetime.now().isoformat(timespec='seconds')} 시작 ===\n")
        sys.stdout = f
        sys.stderr = f
        return log_path
    except Exception:
        return Path()


_LOG_FILE = _init_file_log()


def default_download_dir() -> Path:
    """기본 다운로드 폴더: 내 문서\\팝빌 세금계산서."""
    docs = Path.home() / "Documents"
    if not docs.exists():
        docs = Path.home()
    return docs / "팝빌 세금계산서"


APP_DIR = app_data_dir()
SETTINGS_FILE = APP_DIR / "settings.json"
CACHE_DIR = APP_DIR / "cache"
CACHE_DIR.mkdir(exist_ok=True)


def load_settings() -> dict:
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_settings(data: dict) -> None:
    SETTINGS_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def get_download_dir() -> Path:
    s = load_settings()
    path = s.get("download_dir")
    if path:
        p = Path(path)
        try:
            p.mkdir(parents=True, exist_ok=True)
            return p
        except Exception:
            pass
    p = default_download_dir()
    p.mkdir(parents=True, exist_ok=True)
    return p


def set_download_dir(path: Path) -> None:
    s = load_settings()
    s["download_dir"] = str(path)
    save_settings(s)


# ============================================================
# 거래처 매핑 (사업자번호 → 계정코드)
# ============================================================
VENDOR_MAPPING_FILE = APP_DIR / "vendor_mapping.json"


def _normalize_biznum(num: str) -> str:
    """사업자번호 정규화 (하이픈/공백 제거)."""
    return "".join(c for c in (num or "") if c.isdigit())


def load_vendor_mapping() -> dict:
    if VENDOR_MAPPING_FILE.exists():
        try:
            return json.loads(VENDOR_MAPPING_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_vendor_mapping(mapping: dict) -> None:
    VENDOR_MAPPING_FILE.write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def lookup_vendor(biznum: str) -> dict | None:
    key = _normalize_biznum(biznum)
    if not key:
        return None
    mapping = load_vendor_mapping()
    return mapping.get(key)


def upsert_vendor(biznum: str, account: str,
                  cost_center: str = "", memo: str = "",
                  rules: list | None = None) -> None:
    key = _normalize_biznum(biznum)
    if not key:
        return
    mapping = load_vendor_mapping()
    existing = mapping.get(key, {})
    mapping[key] = {
        "account": account,
        "cost_center": cost_center,
        "memo": memo,
        "rules": rules if rules is not None else existing.get("rules", []),
    }
    save_vendor_mapping(mapping)


def resolve_vendor_fields(info: dict, item_name: str) -> dict:
    """거래처 기본값 + 품목 규칙 병합. 규칙 필드가 비어있으면 기본값 사용."""
    default = {
        "account": info.get("account") or info.get("account_code", ""),
        "cost_center": info.get("cost_center", ""),
        "memo": info.get("memo", ""),
    }
    rules = info.get("rules") or []
    name = item_name or ""
    for rule in rules:
        kw = rule.get("keyword", "").strip()
        if kw and kw in name:
            return {
                "account": rule.get("account") or default["account"],
                "cost_center": rule.get("cost_center") or default["cost_center"],
                "memo": rule.get("memo") or default["memo"],
            }
    return default


def delete_vendor(biznum: str) -> None:
    key = _normalize_biznum(biznum)
    mapping = load_vendor_mapping()
    if key in mapping:
        del mapping[key]
        save_vendor_mapping(mapping)


# ============================================================
# GMS 전표 입력 완료 기록 (NTS 승인번호 기반)
# ============================================================
ENTERED_FILE = APP_DIR / "entered_vouchers.json"


def load_entered() -> set:
    if ENTERED_FILE.exists():
        try:
            data = json.loads(ENTERED_FILE.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return set(data)
        except Exception:
            pass
    return set()


def mark_entered(nts_list: list) -> None:
    current = load_entered()
    current.update(nts_list)
    ENTERED_FILE.write_text(
        json.dumps(sorted(current), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ============================================================
# 거래처 매핑 관리 GUI 창
# ============================================================
class VendorMappingDialog(tk.Toplevel):
    """거래처별 계정코드/적요 매핑 관리 창."""

    def __init__(self, master, invoices: list | None = None,
                 prefill_biznum: str = "",
                 prefill_name: str = "",
                 prefill_item: str = "") -> None:
        super().__init__(master)
        self.title("거래처 매핑 관리")
        self.geometry("1000x760")
        self.transient(master)
        self.grab_set()
        # 현재 조회 결과에서 고유 거래처/품목 목록 생성
        self._all_invoices = invoices or []
        self._vendor_options: list[tuple[str, str]] = []  # (display, biznum)
        self._vendor_items: dict[str, list[str]] = {}     # biznum → item list
        self._build_indexes()
        self._current_rules: list[dict] = []
        self._build()
        self._reload()
        if prefill_biznum:
            self.biznum_var.set(prefill_biznum)
            # 기존 등록 정보 있으면 그 값으로 폼 채우기
            self._load_existing_vendor(prefill_biznum)

    def _build_indexes(self) -> None:
        seen = set()
        for inv in self._all_invoices:
            num = _normalize_biznum(
                inv.get("invoicerCorpNum") or inv.get("supplierCorpNum") or ""
            )
            if not num:
                continue
            name = inv.get("invoicerCorpName") or inv.get("supplierCorpName") or ""
            if num not in seen:
                seen.add(num)
                self._vendor_options.append((f"{name} ({num})", num))
            item = inv.get("itemName") or ""
            if item:
                self._vendor_items.setdefault(num, [])
                if item not in self._vendor_items[num]:
                    self._vendor_items[num].append(item)
        self._vendor_options.sort()

    def _build(self) -> None:
        # 상단 입력
        top = tk.LabelFrame(self, text="매핑 추가/수정", padx=10, pady=8,
                            font=("맑은 고딕", 10, "bold"))
        top.pack(fill="x", padx=10, pady=(10, 5))

        # 목록에서 선택 콤보
        tk.Label(top, text="목록에서 선택:").grid(row=0, column=0, sticky="w", pady=3)
        self.pick_var = tk.StringVar()
        pick_values = [d for d, _ in self._vendor_options]
        self.pick_combo = ttk.Combobox(
            top, textvariable=self.pick_var, width=42,
            values=pick_values,
        )
        self.pick_combo.grid(row=0, column=1, columnspan=3, sticky="w", pady=3)
        self.pick_combo.bind("<<ComboboxSelected>>", self._on_pick_vendor)

        self.biznum_var = tk.StringVar()
        self.account_var = tk.StringVar()
        self.cost_center_var = tk.StringVar()
        self.memo_var = tk.StringVar()

        rows = [
            ("사업자번호:", self.biznum_var, 16),
            ("계정:", self.account_var, 34),
            ("코스트센터:", self.cost_center_var, 34),
            ("적요:", self.memo_var, 34),
        ]
        for i, (label, var, width) in enumerate(rows):
            tk.Label(top, text=label).grid(row=1 + i // 2, column=(i % 2) * 2,
                                            sticky="w", pady=3, padx=(0, 4))
            tk.Entry(top, textvariable=var, width=width).grid(
                row=1 + i // 2, column=(i % 2) * 2 + 1, sticky="w", pady=3, padx=(0, 16)
            )

        tk.Label(top, text="※ 계정/코스트센터는 코드 또는 명을 입력 (GMS 에서 Enter 로 자동완성)",
                 fg="gray").grid(row=3, column=0, columnspan=4, sticky="w", pady=(6, 0))
        tk.Label(top, text="※ 적요: {월}, {년}, {일} 사용 가능 (예: '{월} 유지보수료' → '4월 유지보수료')",
                 fg="gray").grid(row=4, column=0, columnspan=4, sticky="w")

        btn_row = tk.Frame(top)
        btn_row.grid(row=5, column=0, columnspan=4, sticky="w", pady=(8, 0))
        tk.Button(btn_row, text="추가/수정", width=12,
                  bg="#2f80ed", fg="white",
                  command=self.on_upsert).pack(side="left", padx=2)
        tk.Button(btn_row, text="입력 초기화", width=12,
                  command=self.on_clear).pack(side="left", padx=2)

        # 품목 규칙 섹션
        rules_frame = tk.LabelFrame(self, text="품목 규칙 (거래처 기본값 덮어쓰기)",
                                    padx=10, pady=6, font=("맑은 고딕", 10, "bold"))
        rules_frame.pack(fill="x", padx=10, pady=5)

        self.rule_kw_var = tk.StringVar()
        self.rule_acct_var = tk.StringVar()
        self.rule_cc_var = tk.StringVar()
        self.rule_memo_var = tk.StringVar()

        tk.Label(rules_frame, text="품목 포함어:").grid(row=0, column=0, sticky="w", pady=2)
        self.rule_kw_combo = ttk.Combobox(
            rules_frame, textvariable=self.rule_kw_var, width=24,
        )
        self.rule_kw_combo.grid(row=0, column=1, sticky="w", padx=(0, 16), pady=2)

        tk.Label(rules_frame, text="계정:").grid(row=0, column=2, sticky="w", pady=2)
        tk.Entry(rules_frame, textvariable=self.rule_acct_var, width=24).grid(
            row=0, column=3, sticky="w", padx=(0, 16), pady=2)

        tk.Label(rules_frame, text="코스트센터:").grid(row=1, column=0, sticky="w", pady=2)
        tk.Entry(rules_frame, textvariable=self.rule_cc_var, width=24).grid(
            row=1, column=1, sticky="w", padx=(0, 16), pady=2)

        tk.Label(rules_frame, text="적요:").grid(row=1, column=2, sticky="w", pady=2)
        tk.Entry(rules_frame, textvariable=self.rule_memo_var, width=24).grid(
            row=1, column=3, sticky="w", padx=(0, 16), pady=2)

        tk.Label(rules_frame, text="※ 빈 값은 거래처 기본값 상속",
                 fg="gray").grid(row=2, column=0, columnspan=4, sticky="w")

        rule_btns = tk.Frame(rules_frame)
        rule_btns.grid(row=3, column=0, columnspan=4, sticky="w", pady=(4, 0))
        tk.Button(rule_btns, text="규칙 추가", width=10,
                  command=self.on_add_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="규칙 수정", width=10,
                  command=self.on_update_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="규칙 삭제", width=10,
                  command=self.on_delete_rule).pack(side="left", padx=2)
        tk.Button(rule_btns, text="입력 초기화", width=10,
                  command=self.on_clear_rule_form).pack(side="left", padx=2)
        tk.Button(rule_btns, text="거래처+규칙 저장", width=18,
                  bg="#2f80ed", fg="white",
                  command=self.on_upsert).pack(side="left", padx=(16, 2))

        self.rules_tree = ttk.Treeview(
            rules_frame,
            columns=("kw", "acct", "cc", "memo"),
            show="headings",
            height=4,
            selectmode="browse",
        )
        self.rules_tree.bind("<<TreeviewSelect>>", self._on_rule_select)
        for c, t, w in [("kw", "품목 포함어", 160),
                        ("acct", "계정", 200),
                        ("cc", "코스트센터", 160),
                        ("memo", "적요", 260)]:
            self.rules_tree.heading(c, text=t)
            self.rules_tree.column(c, width=w, anchor="w")
        self.rules_tree.grid(row=4, column=0, columnspan=4, sticky="ew", pady=(6, 0))
        rules_frame.columnconfigure(4, weight=1)

        # 표
        mid = tk.LabelFrame(self, text="등록된 매핑", padx=6, pady=6,
                            font=("맑은 고딕", 10, "bold"))
        mid.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("biznum", "account", "cost_center", "memo")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings",
                                 selectmode="browse", height=14)
        headers = {
            "biznum": ("사업자번호", 130, "center"),
            "account": ("계정", 260, "w"),
            "cost_center": ("코스트센터", 220, "w"),
            "memo": ("적요", 240, "w"),
        }
        for c in cols:
            text, w, anchor = headers[c]
            self.tree.heading(c, text=text)
            self.tree.column(c, width=w, anchor=anchor)

        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # 하단
        bot = tk.Frame(self)
        bot.pack(fill="x", padx=10, pady=(5, 10))
        tk.Button(bot, text="선택 삭제", width=12,
                  command=self.on_delete).pack(side="left", padx=2)
        tk.Button(bot, text="파일 열기", width=12,
                  command=self.on_open_file).pack(side="left", padx=2)
        tk.Button(bot, text="닫기", width=12,
                  command=self.destroy).pack(side="right", padx=2)

    def _reload(self) -> None:
        self.tree.delete(*self.tree.get_children())
        mapping = load_vendor_mapping()
        for biznum, info in sorted(mapping.items()):
            rules_count = len(info.get("rules") or [])
            rules_mark = f" (규칙 {rules_count})" if rules_count else ""
            self.tree.insert(
                "", "end", iid=biznum,
                values=(
                    biznum,
                    info.get("account", info.get("account_code", "")) + rules_mark,
                    info.get("cost_center", ""),
                    info.get("memo", ""),
                ),
            )

    def _on_select(self, _evt=None) -> None:
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        self._load_existing_vendor(iid)

    def _load_existing_vendor(self, biznum: str) -> None:
        mapping = load_vendor_mapping()
        info = mapping.get(biznum)
        if not info:
            return
        self.biznum_var.set(biznum)
        self.account_var.set(
            info.get("account") or info.get("account_code", "")
        )
        self.cost_center_var.set(info.get("cost_center", ""))
        self.memo_var.set(info.get("memo", ""))
        self._current_rules = list(info.get("rules") or [])
        self._refresh_rules_tree()
        # 품목 자동완성 후보 업데이트
        items = self._vendor_items.get(biznum, [])
        self.rule_kw_combo["values"] = items

    def _refresh_rules_tree(self) -> None:
        self.rules_tree.delete(*self.rules_tree.get_children())
        for i, r in enumerate(self._current_rules):
            self.rules_tree.insert(
                "", "end", iid=str(i),
                values=(
                    r.get("keyword", ""),
                    r.get("account", ""),
                    r.get("cost_center", ""),
                    r.get("memo", ""),
                ),
            )

    def _on_pick_vendor(self, _evt=None) -> None:
        sel = self.pick_var.get()
        for display, num in self._vendor_options:
            if display == sel:
                self.biznum_var.set(num)
                self._load_existing_vendor(num)
                break

    def _auto_save_vendor(self) -> None:
        """사업자번호가 유효하면 현재 폼 + 규칙을 자동 저장."""
        biznum = self.biznum_var.get().strip()
        if not biznum or not _normalize_biznum(biznum):
            return
        upsert_vendor(
            biznum=biznum,
            account=self.account_var.get().strip(),
            cost_center=self.cost_center_var.get().strip(),
            memo=self.memo_var.get().strip(),
            rules=list(self._current_rules),
        )
        self._reload()
        # 상단 거래처 목록에서 현재 행 유지 선택
        try:
            for iid in self.tree.get_children():
                if iid == _normalize_biznum(biznum):
                    self.tree.selection_set(iid)
                    break
        except Exception:
            pass

    def on_add_rule(self) -> None:
        kw = self.rule_kw_var.get().strip()
        if not kw:
            messagebox.showerror("오류", "품목 포함어를 입력하세요.")
            return
        rule = {
            "keyword": kw,
            "account": self.rule_acct_var.get().strip(),
            "cost_center": self.rule_cc_var.get().strip(),
            "memo": self.rule_memo_var.get().strip(),
        }
        self._current_rules.append(rule)
        self._refresh_rules_tree()
        self.rule_kw_var.set("")
        self.rule_acct_var.set("")
        self.rule_cc_var.set("")
        self.rule_memo_var.set("")
        self._auto_save_vendor()

    def on_delete_rule(self) -> None:
        sel = self.rules_tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        if 0 <= idx < len(self._current_rules):
            del self._current_rules[idx]
            self._refresh_rules_tree()
            self.on_clear_rule_form()
            self._auto_save_vendor()

    def on_update_rule(self) -> None:
        sel = self.rules_tree.selection()
        if not sel:
            messagebox.showinfo("알림", "수정할 규칙을 선택하세요.")
            return
        idx = int(sel[0])
        kw = self.rule_kw_var.get().strip()
        if not kw:
            messagebox.showerror("오류", "품목 포함어를 입력하세요.")
            return
        if 0 <= idx < len(self._current_rules):
            self._current_rules[idx] = {
                "keyword": kw,
                "account": self.rule_acct_var.get().strip(),
                "cost_center": self.rule_cc_var.get().strip(),
                "memo": self.rule_memo_var.get().strip(),
            }
            self._refresh_rules_tree()
            self.on_clear_rule_form()
            self._auto_save_vendor()

    def _on_rule_select(self, _evt=None) -> None:
        sel = self.rules_tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        if 0 <= idx < len(self._current_rules):
            r = self._current_rules[idx]
            self.rule_kw_var.set(r.get("keyword", ""))
            self.rule_acct_var.set(r.get("account", ""))
            self.rule_cc_var.set(r.get("cost_center", ""))
            self.rule_memo_var.set(r.get("memo", ""))

    def on_clear_rule_form(self) -> None:
        self.rule_kw_var.set("")
        self.rule_acct_var.set("")
        self.rule_cc_var.set("")
        self.rule_memo_var.set("")
        try:
            for iid in self.rules_tree.selection():
                self.rules_tree.selection_remove(iid)
        except Exception:
            pass

    def on_upsert(self) -> None:
        biznum = self.biznum_var.get().strip()
        if not biznum:
            messagebox.showerror("오류", "사업자번호를 입력하세요.")
            return
        if not _normalize_biznum(biznum):
            messagebox.showerror("오류", "사업자번호가 숫자가 아닙니다.")
            return
        upsert_vendor(
            biznum=biznum,
            account=self.account_var.get().strip(),
            cost_center=self.cost_center_var.get().strip(),
            memo=self.memo_var.get().strip(),
            rules=list(self._current_rules),
        )
        self._reload()
        self.on_clear()

    def on_clear(self) -> None:
        for v in (self.biznum_var, self.account_var,
                  self.cost_center_var, self.memo_var,
                  self.rule_kw_var, self.rule_acct_var,
                  self.rule_cc_var, self.rule_memo_var,
                  self.pick_var):
            v.set("")
        self._current_rules = []
        self._refresh_rules_tree()

    def on_delete(self) -> None:
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("알림", "삭제할 행을 선택하세요.")
            return
        iid = sel[0]
        if messagebox.askyesno("삭제 확인", f"사업자번호 {iid} 매핑을 삭제합니까?"):
            delete_vendor(iid)
            self._reload()

    def on_open_file(self) -> None:
        try:
            os.startfile(str(VENDOR_MAPPING_FILE))  # type: ignore[attr-defined]
        except Exception as e:
            messagebox.showerror("열기 실패", str(e))


# ============================================================
# 팝빌 래퍼
# ============================================================
def build_service() -> HTTaxinvoiceService:
    svc = HTTaxinvoiceService(LINK_ID, SECRET_KEY)
    svc.IsTest = IS_TEST
    return svc


def obj_to_dict(obj) -> dict:
    d: dict = {}
    for k in dir(obj):
        if k.startswith("_"):
            continue
        v = getattr(obj, k, None)
        if callable(v):
            continue
        try:
            json.dumps(v)
            d[k] = v
        except (TypeError, ValueError):
            d[k] = str(v)
    return d


def cache_path(sdate: str, edate: str) -> Path:
    return CACHE_DIR / f"invoices_{sdate}_{edate}.json"


def load_cache(sdate: str, edate: str) -> list[dict] | None:
    p = cache_path(sdate, edate)
    if not p.exists():
        return None
    age = (time.time() - p.stat().st_mtime) / 60
    if age > CACHE_TTL_MIN:
        return None
    return json.loads(p.read_text(encoding="utf-8"))


def save_cache(sdate: str, edate: str, items: list[dict]) -> None:
    cache_path(sdate, edate).write_text(
        json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def fetch_invoices(
    svc: HTTaxinvoiceService,
    sdate: str,
    edate: str,
    log: callable,
) -> list[dict]:
    """캐시 → 실제 수집 순. log(msg) 로 진행 상황 알림."""
    cached = load_cache(sdate, edate)
    if cached is not None:
        log(f"캐시 재사용: {len(cached)}건")
        return cached

    log("인증서 확인 중...")
    svc.checkCertValidation(CORP_NUM, USER_ID)

    log("수집 작업 요청 중...")
    job_id = None
    last_err = None
    for attempt in range(3):
        try:
            job_id = svc.requestJob(
                CorpNum=CORP_NUM, Type=INVOICE_TYPE,
                DType="I",  # I = 발행일자, W = 작성일자
                SDate=sdate, EDate=edate, UserID=USER_ID,
            )
            break
        except PopbillException as e:
            last_err = e
            log(f"requestJob 실패 {attempt+1}/3: {e.message}")
            time.sleep(5)
    if job_id is None:
        raise last_err or RuntimeError("requestJob 실패")
    log(f"JobID={job_id}")

    for i in range(60):
        st_obj = svc.getJobState(CORP_NUM, job_id, USER_ID)
        st = st_obj.jobState
        err = getattr(st_obj, "errorCode", 0) or 0
        reason = getattr(st_obj, "errorReason", "") or ""
        cnt = getattr(st_obj, "collectCount", 0) or 0
        log(f"[{i+1}] 상태={st} 수집={cnt}건 {reason}")
        if st in (2, 3):
            if err == 1 or "완료" in reason:
                log(f"수집 완료: {cnt}건")
                break
            raise RuntimeError(f"수집 실패: {reason}")
        time.sleep(5)
    else:
        raise TimeoutError("수집 5분 초과")

    log("목록 조회 중...")
    items: list[dict] = []
    page = 1
    while True:
        for attempt in range(10):
            try:
                res = svc.search(
                    CorpNum=CORP_NUM, JobID=job_id,
                    Type=["N", "M"], TaxType=["T", "N", "Z"],
                    PurposeType=["R", "C", "N"],
                    TaxRegIDType="", TaxRegIDYN="", TaxRegID="",
                    Page=page, PerPage=500, Order="D", UserID=USER_ID,
                )
                break
            except PopbillException as e:
                if e.code == -11110005 and attempt < 9:
                    time.sleep(3)
                    continue
                raise
        lst = getattr(res, "list", []) or []
        total = getattr(res, "total", 0) or 0
        items.extend(obj_to_dict(x) for x in lst)
        log(f"page {page}: {len(items)}/{total}")
        if len(items) >= total or not lst:
            break
        page += 1

    save_cache(sdate, edate, items)
    return items


# ============================================================
# 값 헬퍼
# ============================================================
def g(d: dict, *keys, default=""):
    for k in keys:
        v = d.get(k)
        if v not in (None, ""):
            return v
    return default


def fmt_amount(v) -> str:
    try:
        return f"{int(v):,}"
    except (ValueError, TypeError):
        return str(v) if v else ""


def fmt_date(s: str) -> str:
    if not s or len(s) != 8:
        return s
    return f"{s[:4]}-{s[4:6]}-{s[6:]}"


# ============================================================
# PDF 다운로드 (asyncio 이벤트 루프 내에서)
# ============================================================
async def download_selected_items(
    svc: HTTaxinvoiceService,
    items: list[dict],
    progress: callable,
    out_dir: Path,
) -> tuple[int, int]:
    from playwright.async_api import async_playwright

    ok = 0
    fail = 0
    total = len(items)
    out_dir.mkdir(parents=True, exist_ok=True)
    async with async_playwright() as p:
        # Windows 기본 설치된 Edge 사용 (exe 배포 시 Chromium 번들 불필요)
        try:
            browser = await p.chromium.launch(headless=True, channel="msedge")
        except Exception:
            browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        for n, inv in enumerate(items, 1):
            nts = g(inv, "ntsconfirmNum", "ntsConfirmNum")
            progress(n, total, nts)
            if not nts:
                fail += 1
                continue
            try:
                url = svc.getPrintURL(CORP_NUM, nts, USER_ID)
            except PopbillException:
                fail += 1
                continue
            out = out_dir / f"{nts}.pdf"
            try:
                page = await context.new_page()
                await page.goto(url, wait_until="networkidle", timeout=60000)
                await page.pdf(
                    path=str(out), format="A4", print_background=True,
                    margin={"top": "10mm", "right": "10mm",
                            "bottom": "10mm", "left": "10mm"},
                )
                await page.close()
                ok += 1
            except Exception:
                fail += 1
        await browser.close()
    return ok, fail


# ============================================================
# GUI 앱
# ============================================================
class PopbillApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"AI 세금계산서  v{APP_VERSION}")
        self.geometry("1320x800")
        self.minsize(1000, 600)
        # 아이콘 적용 (PyInstaller 번들 또는 소스 실행 모두 대응)
        try:
            if hasattr(sys, "_MEIPASS"):
                icon_path = Path(sys._MEIPASS) / "app.ico"
            else:
                icon_path = Path(__file__).parent / "app.ico"
            if icon_path.exists():
                self.iconbitmap(str(icon_path))
        except Exception:
            pass

        self.svc = build_service()
        self.items: list[dict] = []           # 전체 조회 결과
        self.filtered: list[dict] = []        # 현재 필터 적용 후
        self.checked: set[str] = set()        # 체크된 승인번호(iid) 집합
        self.entered_set: set[str] = load_entered()  # GMS 입력 완료 NTS 집합
        self.focused_nts: str = ""            # 마지막 클릭된 행의 NTS
        self._busy = False
        self._settings = load_settings()

        self._build_ui()

    # ---------- UI 구성 ----------
    def _build_ui(self) -> None:
        # 상단: 조회 조건
        top = tk.LabelFrame(self, text="조회 조건", padx=10, pady=8, font=("맑은 고딕", 10, "bold"))
        top.pack(fill="x", padx=10, pady=(10, 5))

        tk.Label(top, text="대상 사업자:").grid(row=0, column=0, sticky="w")
        tk.Label(top, text=CORP_NUM, font=("맑은 고딕", 10, "bold"),
                 fg="#2f80ed").grid(row=0, column=1, sticky="w", padx=(4, 20))

        tk.Label(top, text="시작일:").grid(row=0, column=2, sticky="w")
        self.start_var = tk.StringVar()
        tk.Entry(top, textvariable=self.start_var, width=12).grid(row=0, column=3, padx=4)

        tk.Label(top, text="종료일:").grid(row=0, column=4, sticky="w")
        self.end_var = tk.StringVar()
        tk.Entry(top, textvariable=self.end_var, width=12).grid(row=0, column=5, padx=4)

        today = datetime.now()
        # 저장된 마지막 조회 조건 복원 (없으면 이번 달)
        saved_start = self._settings.get("last_start_date", "")
        saved_end = self._settings.get("last_end_date", "")
        self.start_var.set(saved_start or today.replace(day=1).strftime("%Y-%m-%d"))
        self.end_var.set(saved_end or today.strftime("%Y-%m-%d"))

        tk.Label(top, text="(YYYY-MM-DD, 최대 3개월)",
                 fg="gray").grid(row=0, column=6, padx=4)

        self.search_btn = tk.Button(top, text="조회", width=10,
                                    bg="#2f80ed", fg="white",
                                    font=("맑은 고딕", 10, "bold"),
                                    command=self.on_search)
        self.search_btn.grid(row=0, column=7, padx=(20, 4))

        # 필터 영역
        flt = tk.LabelFrame(self, text="필터", padx=10, pady=6,
                            font=("맑은 고딕", 10, "bold"))
        flt.pack(fill="x", padx=10, pady=(0, 5))

        tk.Label(flt, text="거래처/사업자번호:").grid(row=0, column=0, sticky="w")
        self.filter_keyword = tk.StringVar(
            value=self._settings.get("last_filter_keyword", "")
        )
        self.filter_keyword.trace_add(
            "write", lambda *_: (self._save_filters(), self._apply_filter())
        )
        tk.Entry(flt, textvariable=self.filter_keyword, width=28).grid(row=0, column=1, padx=4)

        tk.Label(flt, text="  과세구분:").grid(row=0, column=2, sticky="w")
        self.filter_taxtype = tk.StringVar(
            value=self._settings.get("last_filter_taxtype", "전체")
        )
        ttk.Combobox(
            flt, textvariable=self.filter_taxtype, width=8, state="readonly",
            values=["전체", "과세", "면세", "영세"],
        ).grid(row=0, column=3, padx=4)
        self.filter_taxtype.trace_add(
            "write", lambda *_: (self._save_filters(), self._apply_filter())
        )

        tk.Label(flt, text="  문서유형:").grid(row=0, column=4, sticky="w")
        self.filter_doctype = tk.StringVar(
            value=self._settings.get("last_filter_doctype", "전체")
        )
        ttk.Combobox(
            flt, textvariable=self.filter_doctype, width=8, state="readonly",
            values=["전체", "일반", "수정"],
        ).grid(row=0, column=5, padx=4)
        self.filter_doctype.trace_add(
            "write", lambda *_: (self._save_filters(), self._apply_filter())
        )

        tk.Label(flt, text="  전표:").grid(row=0, column=6, sticky="w")
        self.filter_entered = tk.StringVar(
            value=self._settings.get("last_filter_entered", "전체")
        )
        ttk.Combobox(
            flt, textvariable=self.filter_entered, width=8, state="readonly",
            values=["전체", "미입력", "입력완료"],
        ).grid(row=0, column=7, padx=4)
        self.filter_entered.trace_add(
            "write", lambda *_: (self._save_filters(), self._apply_filter())
        )

        tk.Label(flt, text="  매핑:").grid(row=0, column=8, sticky="w")
        self.filter_mapped = tk.StringVar(
            value=self._settings.get("last_filter_mapped", "전체")
        )
        ttk.Combobox(
            flt, textvariable=self.filter_mapped, width=8, state="readonly",
            values=["전체", "미매핑", "매핑완료"],
        ).grid(row=0, column=9, padx=4)
        self.filter_mapped.trace_add(
            "write", lambda *_: (self._save_filters(), self._apply_filter())
        )

        tk.Button(flt, text="필터 초기화", command=self._reset_filter,
                  width=12).grid(row=0, column=10, padx=(20, 4))

        self.filter_stat = tk.Label(flt, text="", fg="gray")
        self.filter_stat.grid(row=0, column=11, padx=10, sticky="w")

        # 중단: 표
        mid = tk.Frame(self)
        mid.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("chk", "num", "map", "entered", "wdate", "idate",
                "supplier", "supid", "doc", "taxtype", "item", "remark",
                "supply", "tax", "total", "nts")
        headers = {
            "chk": ("선택", 50, "center"),
            "num": ("#", 45, "center"),
            "map": ("매핑", 50, "center"),
            "entered": ("전표", 55, "center"),
            "wdate": ("작성일자", 90, "center"),
            "idate": ("발행일자", 90, "center"),
            "supplier": ("공급자", 220, "w"),
            "supid": ("사업자번호", 110, "center"),
            "doc": ("유형", 45, "center"),
            "taxtype": ("과세", 45, "center"),
            "item": ("품목(대표)", 160, "w"),
            "remark": ("비고", 160, "w"),
            "supply": ("공급가액", 105, "e"),
            "tax": ("세액", 85, "e"),
            "total": ("합계금액", 105, "e"),
            "nts": ("승인번호", 200, "center"),
        }
        self.tree = ttk.Treeview(mid, columns=cols, show="headings",
                                 selectmode="none", height=25)
        for c in cols:
            text, w, anchor = headers[c]
            self.tree.heading(c, text=text)
            self.tree.column(c, width=w, minwidth=w, anchor=anchor, stretch=False)

        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        mid.rowconfigure(0, weight=1)
        mid.columnconfigure(0, weight=1)

        self.tree.tag_configure("odd", background="#f7f7f7")
        self.tree.tag_configure("even", background="#ffffff")
        self.tree.tag_configure(
            "checked", background="#cde4ff", foreground="#0b3d91"
        )
        self.tree.tag_configure("negative", foreground="#d32f2f")
        self.tree.tag_configure(
            "checked_neg", background="#cde4ff", foreground="#d32f2f"
        )
        # 전표 입력 완료 행 (연한 녹색 배경)
        self.tree.tag_configure("entered", background="#e8f5e9")
        self.tree.tag_configure(
            "entered_neg", background="#e8f5e9", foreground="#d32f2f"
        )

        # 클릭으로 체크 토글
        self.tree.bind("<Button-1>", self._on_tree_click)

        # 저장 경로 표시/선택
        path_frame = tk.Frame(self)
        path_frame.pack(fill="x", padx=10, pady=(5, 0))
        tk.Label(path_frame, text="저장 폴더:",
                 font=("맑은 고딕", 9)).pack(side="left")
        self.path_var = tk.StringVar(value=str(get_download_dir()))
        tk.Entry(path_frame, textvariable=self.path_var,
                 state="readonly", width=70).pack(side="left", fill="x",
                                                   expand=True, padx=4)
        tk.Button(path_frame, text="폴더 선택", width=10,
                  command=self.on_pick_folder).pack(side="left", padx=2)
        tk.Button(path_frame, text="열기", width=6,
                  command=self.on_open_folder).pack(side="left", padx=2)

        # GMS 옵션
        opt_frame = tk.Frame(self)
        opt_frame.pack(fill="x", padx=10, pady=(4, 0))
        s = load_settings()
        self.headless_var = tk.BooleanVar(value=s.get("gms_headless", True))
        tk.Checkbutton(
            opt_frame,
            text="GMS 백그라운드 실행 (창 안 보임, PC 사용 가능)",
            variable=self.headless_var,
            command=self._save_gms_options,
            font=("맑은 고딕", 9),
        ).pack(side="left")
        self.auto_close_var = tk.BooleanVar(value=s.get("gms_auto_close", True))
        tk.Checkbutton(
            opt_frame,
            text="완료 후 GMS 창 자동 종료",
            variable=self.auto_close_var,
            command=self._save_gms_options,
            font=("맑은 고딕", 9),
        ).pack(side="left", padx=(16, 0))

        # 하단: 상태/버튼
        bot = tk.Frame(self)
        bot.pack(fill="x", padx=10, pady=(5, 10))

        self.status_var = tk.StringVar(value="대기")
        tk.Label(bot, textvariable=self.status_var, anchor="w",
                 fg="#333").pack(side="left", fill="x", expand=True)

        tk.Label(bot, text=f"v{APP_VERSION}", fg="#999",
                 font=("맑은 고딕", 8)).pack(side="left", padx=(0, 8))

        self.count_var = tk.StringVar(value="선택 0건")
        tk.Label(bot, textvariable=self.count_var,
                 font=("맑은 고딕", 10, "bold")).pack(side="right", padx=10)

        tk.Button(bot, text="전체 선택", width=10,
                  command=self.on_check_all).pack(side="left", padx=2)
        tk.Button(bot, text="전체 해제", width=10,
                  command=self.on_uncheck_all).pack(side="left", padx=2)
        tk.Button(bot, text="거래처 매핑", width=12,
                  command=self.on_vendor_mapping).pack(side="left", padx=(16, 2))
        self.voucher_btn = tk.Button(bot, text="GMS 전표 입력", width=14,
                                     bg="#d97706", fg="white",
                                     font=("맑은 고딕", 10, "bold"),
                                     command=self.on_gms_entry,
                                     state="disabled")
        self.voucher_btn.pack(side="right", padx=4)
        self.save_btn = tk.Button(bot, text="저장(PDF)", width=12,
                                  bg="#27ae60", fg="white",
                                  font=("맑은 고딕", 10, "bold"),
                                  command=self.on_save, state="disabled")
        self.save_btn.pack(side="right", padx=4)
        self.excel_btn = tk.Button(bot, text="엑셀 다운로드", width=12,
                                   bg="#217346", fg="white",
                                   font=("맑은 고딕", 10, "bold"),
                                   command=self.on_export_excel, state="disabled")
        self.excel_btn.pack(side="right", padx=4)
        tk.Button(bot, text="닫기", width=10,
                  command=self.destroy).pack(side="right", padx=4)

    # ---------- 저장 폴더 ----------
    def on_pick_folder(self) -> None:
        current = self.path_var.get()
        chosen = filedialog.askdirectory(
            title="PDF 저장 폴더 선택",
            initialdir=current if Path(current).exists() else str(Path.home()),
        )
        if chosen:
            p = Path(chosen)
            set_download_dir(p)
            self.path_var.set(str(p))

    def on_open_folder(self) -> None:
        p = Path(self.path_var.get())
        p.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(str(p))  # type: ignore[attr-defined]
        except Exception as e:
            messagebox.showerror("열기 실패", str(e))

    # ---------- 진행 상태 ----------
    def set_status(self, msg: str) -> None:
        self.status_var.set(msg)
        self.update_idletasks()

    def _set_busy(self, busy: bool) -> None:
        self._busy = busy
        self.search_btn.config(state="disabled" if busy else "normal")
        self.save_btn.config(
            state="disabled" if busy or not self.items else "normal"
        )
        self.voucher_btn.config(
            state="disabled" if busy or not self.items else "normal"
        )
        self.excel_btn.config(
            state="disabled" if busy or not self.items else "normal"
        )

    def on_vendor_mapping(self) -> None:
        # 포커스된 행(또는 체크된 첫 행) 에서 사업자번호/공급자명 가져오기
        biznum = ""
        name = ""
        item = ""
        target_nts = self.focused_nts or (
            next(iter(self.checked), "") if self.checked else ""
        )
        if target_nts:
            for inv in self.items:
                if g(inv, "ntsconfirmNum", "ntsConfirmNum") == target_nts:
                    biznum = _normalize_biznum(
                        g(inv, "invoicerCorpNum", "supplierCorpNum")
                    )
                    name = g(inv, "invoicerCorpName", "supplierCorpName")
                    item = g(inv, "itemName")
                    break
        dlg = VendorMappingDialog(
            self,
            invoices=self.items,
            prefill_biznum=biznum,
            prefill_name=name,
            prefill_item=item,
        )
        # 다이얼로그 닫힌 뒤 그리드 재표시 (매핑 표시 갱신)
        self.wait_window(dlg)
        if self.items:
            self._apply_filter()

    def _save_gms_options(self) -> None:
        s = load_settings()
        s["gms_headless"] = bool(self.headless_var.get())
        s["gms_auto_close"] = bool(self.auto_close_var.get())
        save_settings(s)

    def _save_filters(self) -> None:
        s = load_settings()
        s["last_filter_keyword"] = self.filter_keyword.get()
        s["last_filter_taxtype"] = self.filter_taxtype.get()
        s["last_filter_doctype"] = self.filter_doctype.get()
        s["last_filter_entered"] = self.filter_entered.get()
        s["last_filter_mapped"] = self.filter_mapped.get()
        save_settings(s)
        self._settings = s

    def _save_date_range(self, sdate: str, edate: str) -> None:
        s = load_settings()
        s["last_start_date"] = sdate
        s["last_end_date"] = edate
        save_settings(s)
        self._settings = s

    # ---------- GMS 전표 입력 ----------
    def on_gms_entry(self) -> None:
        if self._busy:
            return
        if not self.checked:
            messagebox.showinfo("알림", "선택된 항목이 없습니다.")
            return
        nts_set = set(self.checked)
        selected_items = [
            inv for inv in self.items
            if g(inv, "ntsconfirmNum", "ntsConfirmNum") in nts_set
        ]
        mapping = load_vendor_mapping()

        # 사전 검사: 매핑 없는 거래처 찾기
        missing = []
        valid = []
        for inv in selected_items:
            biznum = _normalize_biznum(g(inv, "invoicerCorpNum", "supplierCorpNum"))
            if not biznum or biznum not in mapping:
                missing.append(inv)
            else:
                valid.append(inv)

        msg = (f"선택 {len(selected_items)}건 중\n"
               f"  ▶ 자동 입력 가능: {len(valid)}건\n"
               f"  ▶ 매핑 없어 스킵: {len(missing)}건\n\n"
               f"GMS 브라우저가 열립니다. 계속하시겠습니까?")
        if not messagebox.askyesno("GMS 전표 입력", msg):
            return

        out_dir = Path(self.path_var.get())
        self._set_busy(True)
        self.set_status("GMS 전표 입력 시작...")
        threading.Thread(
            target=self._gms_thread,
            args=(selected_items, mapping, out_dir),
            daemon=True,
        ).start()

    def _gms_thread(self, items: list[dict], mapping: dict, out_dir: Path) -> None:
        import traceback
        from gms_automation import enter_vouchers

        def log(m: str) -> None:
            self.after(0, self.set_status, m)
            print(m)

        def dl_progress(n: int, total: int, nts: str) -> None:
            self.after(0, self.set_status, f"PDF 다운로드 {n}/{total}: {nts}")

        def gms_progress(n: int, total: int, nts: str) -> None:
            self.after(0, self.set_status, f"GMS 입력 {n}/{total}: {nts}")

        try:
            out_dir.mkdir(parents=True, exist_ok=True)

            # 1단계: 매핑이 있는 건에 대해서만 PDF 사전 다운로드
            need_pdf = []
            for inv in items:
                biznum = _normalize_biznum(g(inv, "invoicerCorpNum", "supplierCorpNum"))
                if not biznum or biznum not in mapping:
                    continue  # 스킵 대상은 PDF 도 불필요
                nts = g(inv, "ntsconfirmNum", "ntsConfirmNum")
                if not nts:
                    continue
                pdf = out_dir / f"{nts}.pdf"
                if not pdf.exists():
                    need_pdf.append(inv)

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            if need_pdf:
                log(f"PDF 다운로드 시작: {len(need_pdf)}건")
                loop.run_until_complete(
                    download_selected_items(self.svc, need_pdf, dl_progress, out_dir)
                )
                log("PDF 다운로드 완료")
            else:
                log("PDF 캐시 사용 (이미 존재)")

            # 2단계: GMS 자동 입력
            log("GMS 전표 입력 시작")
            headless = bool(self.headless_var.get())
            keep_open = not bool(self.auto_close_var.get())
            result = loop.run_until_complete(
                enter_vouchers(
                    items, mapping, out_dir,
                    log=log, progress=gms_progress,
                    headless=headless, keep_open=keep_open,
                )
            )
            loop.close()
            self.after(0, self._on_gms_done, result)
        except Exception as e:
            tb = traceback.format_exc()
            print(tb)  # 파일 로그에 기록
            self.after(0, self._on_gms_error, f"{e}\n\n상세:\n{tb}")

    def _on_gms_done(self, result: dict) -> None:
        self._set_busy(False)
        ok = result.get("ok", 0)
        fail = result.get("fail", 0)
        missing = result.get("missing_vendors", [])
        success_nts = result.get("success_nts", [])
        # 성공한 NTS 를 영구 저장 + 메모리 셋에 반영
        if success_nts:
            mark_entered(success_nts)
            self.entered_set.update(success_nts)
            # 현재 표시 중인 행들의 상태 업데이트
            self._apply_filter()

        lines = [
            f"GMS 전표 입력 완료",
            f"  성공: {ok}건",
            f"  실패: {fail}건",
            f"  매핑 없어 스킵: {len(missing)}건",
        ]
        if missing:
            lines.append("")
            lines.append("매핑 등록이 필요한 거래처:")
            for m in missing[:20]:
                lines.append(
                    f"  - {m.get('vendor_name','')} ({m.get('biznum','')})"
                )
            if len(missing) > 20:
                lines.append(f"  ... 외 {len(missing) - 20}건")

        msg = "\n".join(lines)
        self.set_status(f"완료: 성공 {ok} 실패 {fail} 스킵 {len(missing)}")
        messagebox.showinfo("GMS 전표 입력 결과", msg)

    def _on_gms_error(self, msg: str) -> None:
        self._set_busy(False)
        self.set_status(f"GMS 오류: {msg}")
        messagebox.showerror("GMS 전표 입력 실패", msg)

    # ---------- 엑셀 다운로드 ----------
    def on_export_excel(self) -> None:
        if not HAS_OPENPYXL:
            messagebox.showerror("오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl")
            return
        data = self.filtered if self.filtered else self.items
        if not data:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

        default_name = f"세금계산서_{self.start_var.get()}_{self.end_var.get()}.xlsx"
        path = filedialog.asksaveasfilename(
            title="엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=default_name,
            initialdir=str(Path(self.path_var.get())),
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "세금계산서"

        # 헤더
        excel_headers = [
            "#", "작성일자", "발행일자", "공급자", "사업자번호",
            "유형", "과세구분", "품목(대표)", "비고",
            "공급가액", "세액", "합계금액", "승인번호",
        ]
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        for col_idx, h in enumerate(excel_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 데이터 행
        for i, inv in enumerate(data, 1):
            wdate = fmt_date(g(inv, "writeDate"))
            idate = fmt_date(g(inv, "issueDate", "issuedDate"))
            supplier = g(inv, "invoicerCorpName", "supplierCorpName")
            sup_id = g(inv, "invoicerCorpNum", "supplierCorpNum")
            modify_yn = inv.get("modifyYN")
            doc_label = "수정" if modify_yn is True or str(modify_yn).lower() == "true" else "일반"
            taxtype_code = g(inv, "taxType")
            taxtype_label = self.TAXTYPE_LABEL.get(taxtype_code, taxtype_code or "")
            item_name = g(inv, "itemName", "itemName1")
            remark = g(inv, "remark1", "remark")
            try:
                supply_val = int(float(str(g(inv, "supplyCostTotal", default=0)).replace(",", "") or 0))
            except (ValueError, TypeError):
                supply_val = 0
            try:
                tax_val = int(float(str(g(inv, "taxTotal", default=0)).replace(",", "") or 0))
            except (ValueError, TypeError):
                tax_val = 0
            try:
                total_val = int(float(str(g(inv, "totalAmount", "amountTotal", default=0)).replace(",", "") or 0))
            except (ValueError, TypeError):
                total_val = 0
            nts = g(inv, "ntsconfirmNum", "ntsConfirmNum")

            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=wdate)
            ws.cell(row=row, column=3, value=idate)
            ws.cell(row=row, column=4, value=supplier)
            ws.cell(row=row, column=5, value=sup_id)
            ws.cell(row=row, column=6, value=doc_label)
            ws.cell(row=row, column=7, value=taxtype_label)
            ws.cell(row=row, column=8, value=item_name)
            ws.cell(row=row, column=9, value=remark)
            c_supply = ws.cell(row=row, column=10, value=supply_val)
            c_supply.number_format = '#,##0'
            c_tax = ws.cell(row=row, column=11, value=tax_val)
            c_tax.number_format = '#,##0'
            c_total = ws.cell(row=row, column=12, value=total_val)
            c_total.number_format = '#,##0'
            ws.cell(row=row, column=13, value=nts)

        # 열 너비 조정
        col_widths = [6, 12, 12, 25, 15, 8, 10, 20, 20, 14, 12, 14, 28]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

        # 자동 필터
        ws.auto_filter.ref = ws.dimensions

        try:
            wb.save(path)
            messagebox.showinfo("완료", f"엑셀 저장 완료\n{path}")
        except PermissionError:
            messagebox.showerror("오류", "파일이 이미 열려 있습니다. 닫고 다시 시도하세요.")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def _refresh_count(self) -> None:
        self.count_var.set(f"선택 {len(self.checked)}건 / 전체 {len(self.items)}건")

    # ---------- 조회 ----------
    def on_search(self) -> None:
        if self._busy:
            return
        try:
            sdate = self.start_var.get().replace("-", "").strip()
            edate = self.end_var.get().replace("-", "").strip()
            if len(sdate) != 8 or len(edate) != 8:
                raise ValueError("YYYY-MM-DD 형식으로 입력하세요.")
            datetime.strptime(sdate, "%Y%m%d")
            datetime.strptime(edate, "%Y%m%d")
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        # 날짜 설정 저장 (사람이 입력한 형태 그대로)
        self._save_date_range(
            self.start_var.get().strip(),
            self.end_var.get().strip(),
        )
        self._set_busy(True)
        self.set_status(f"{sdate} ~ {edate} 조회 중...")
        threading.Thread(
            target=self._search_thread, args=(sdate, edate), daemon=True
        ).start()

    def _search_thread(self, sdate: str, edate: str) -> None:
        try:
            items = fetch_invoices(
                self.svc, sdate, edate,
                log=lambda m: self.after(0, self.set_status, m),
            )
            self.after(0, self._on_search_done, items)
        except PopbillException as e:
            self.after(0, self._on_search_error, f"[팝빌] {e.code} {e.message}")
        except Exception as e:
            self.after(0, self._on_search_error, str(e))

    def _on_search_done(self, items: list[dict]) -> None:
        self.items = items
        self.checked.clear()
        self._apply_filter()
        self._set_busy(False)
        self.set_status(f"조회 완료: {len(items)}건")

    def _on_search_error(self, msg: str) -> None:
        self._set_busy(False)
        self.set_status(f"오류: {msg}")
        messagebox.showerror("조회 실패", msg)

    # ---------- 필터 ----------
    TAXTYPE_LABEL = {"T": "과세", "N": "면세", "Z": "영세"}
    DOCTYPE_LABEL = {"N": "일반", "M": "수정"}

    def _reset_filter(self) -> None:
        self.filter_keyword.set("")
        self.filter_taxtype.set("전체")
        self.filter_doctype.set("전체")
        self.filter_entered.set("전체")
        self.filter_mapped.set("전체")

    def _apply_filter(self) -> None:
        keyword = self.filter_keyword.get().strip().lower()
        tt = self.filter_taxtype.get()
        dt = self.filter_doctype.get()
        mapping_keys = set(load_vendor_mapping().keys())

        def match(inv: dict) -> bool:
            if keyword:
                name = (g(inv, "invoicerCorpName", "supplierCorpName") or "").lower()
                num = (g(inv, "invoicerCorpNum", "supplierCorpNum") or "").lower()
                if keyword not in name and keyword not in num.replace("-", ""):
                    return False
            if tt != "전체":
                t = g(inv, "taxType")
                if self.TAXTYPE_LABEL.get(t, "") != tt:
                    return False
            if dt != "전체":
                modify_yn = inv.get("modifyYN")
                is_mod = (modify_yn is True
                          or str(modify_yn).lower() == "true")
                if dt == "수정" and not is_mod:
                    return False
                if dt == "일반" and is_mod:
                    return False
            # 전표 입력 여부 필터
            ef = self.filter_entered.get()
            if ef != "전체":
                nts = g(inv, "ntsconfirmNum", "ntsConfirmNum") or ""
                is_ent = nts in self.entered_set
                if ef == "입력완료" and not is_ent:
                    return False
                if ef == "미입력" and is_ent:
                    return False
            # 매핑 여부 필터
            mf = self.filter_mapped.get()
            if mf != "전체":
                bnum = _normalize_biznum(
                    g(inv, "invoicerCorpNum", "supplierCorpNum")
                )
                is_mapped = bnum in mapping_keys
                if mf == "매핑완료" and not is_mapped:
                    return False
                if mf == "미매핑" and is_mapped:
                    return False
            return True

        self.filtered = [inv for inv in self.items if match(inv)]
        self._populate_tree()
        self._refresh_count()
        self.filter_stat.config(
            text=f"필터 적용: {len(self.filtered)} / {len(self.items)}건"
        )

    def _row_tags(self, idx: int, is_checked: bool,
                  is_negative: bool = False,
                  is_entered: bool = False) -> tuple[str, ...]:
        stripe = "odd" if idx % 2 else "even"
        if is_checked:
            return ("checked_neg",) if is_negative else ("checked",)
        if is_entered:
            return ("entered_neg",) if is_negative else ("entered",)
        if is_negative:
            return (stripe, "negative")
        return (stripe,)

    # ---------- 표 채우기 ----------
    def _populate_tree(self) -> None:
        self.tree.delete(*self.tree.get_children())
        mapping_keys = set(load_vendor_mapping().keys())
        for i, inv in enumerate(self.filtered):
            wdate = fmt_date(g(inv, "writeDate"))
            idate = fmt_date(g(inv, "issueDate", "issuedDate"))
            supplier = g(inv, "invoicerCorpName", "supplierCorpName")
            sup_id = g(inv, "invoicerCorpNum", "supplierCorpNum")
            # 팝빌 modifyYN: True(수정) / False(일반)
            modify_yn = inv.get("modifyYN")
            if modify_yn is True or str(modify_yn).lower() == "true":
                doc_label = "수정"
            else:
                doc_label = "일반"
            taxtype_code = g(inv, "taxType")
            taxtype_label = self.TAXTYPE_LABEL.get(taxtype_code, taxtype_code or "")
            supply_raw = g(inv, "supplyCostTotal", default=0)
            try:
                is_negative = int(float(str(supply_raw).replace(",", "") or 0)) < 0
            except (ValueError, TypeError):
                is_negative = False
            supply = fmt_amount(supply_raw)
            tax = fmt_amount(g(inv, "taxTotal", default=0))
            total = fmt_amount(g(inv, "totalAmount", "amountTotal", default=0))
            nts = g(inv, "ntsconfirmNum", "ntsConfirmNum") or f"_NONTS_{i}"
            item_name = g(inv, "itemName", "itemName1")
            remark = g(inv, "remark1", "remark")
            biznum_norm = _normalize_biznum(sup_id)
            is_mapped = biznum_norm in mapping_keys
            is_checked = nts in self.checked
            is_entered = nts in self.entered_set
            mark = "☑" if is_checked else "☐"
            map_mark = "✓" if is_mapped else ""
            entered_mark = "✓" if is_entered else ""
            self.tree.insert(
                "", "end", iid=nts,
                values=(mark, i + 1, map_mark, entered_mark, wdate, idate,
                        supplier, sup_id, doc_label, taxtype_label,
                        item_name, remark, supply, tax, total, nts),
                tags=self._row_tags(i, is_checked, is_negative, is_entered),
            )

    # ---------- 체크박스 토글 ----------
    def _on_tree_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        if not item:
            return
        self.focused_nts = item
        self._toggle(item)

    def _row_index(self, iid: str) -> int:
        """트리에서 행의 위치 인덱스 (스트라이프 색상 복원용)."""
        try:
            return self.tree.index(iid)
        except Exception:
            return 0

    def _apply_row_style(self, iid: str, is_checked: bool) -> None:
        # 기존 tags 에서 negative 여부 유지
        current = self.tree.item(iid, "tags") or ()
        is_neg = "negative" in current or "checked_neg" in current
        self.tree.item(
            iid,
            tags=self._row_tags(self._row_index(iid), is_checked, is_neg),
        )
        self.tree.set(iid, "chk", "☑" if is_checked else "☐")

    def _toggle(self, iid: str) -> None:
        if iid in self.checked:
            self.checked.remove(iid)
            self._apply_row_style(iid, False)
        else:
            self.checked.add(iid)
            self._apply_row_style(iid, True)
        self._refresh_count()

    def on_check_all(self) -> None:
        """현재 필터에 보이는 행만 전체 체크."""
        for iid in self.tree.get_children():
            if iid not in self.checked:
                self.checked.add(iid)
                self._apply_row_style(iid, True)
        self._refresh_count()

    def on_uncheck_all(self) -> None:
        """현재 필터에 보이는 행만 전체 해제."""
        for iid in self.tree.get_children():
            if iid in self.checked:
                self.checked.remove(iid)
                self._apply_row_style(iid, False)
        self._refresh_count()

    # ---------- 저장 ----------
    def on_save(self) -> None:
        if self._busy:
            return
        if not self.checked:
            messagebox.showinfo("알림", "선택된 항목이 없습니다.")
            return
        # 체크된 승인번호에 해당하는 원본 invoice 리스트
        nts_set = set(self.checked)
        selected_items = [
            inv for inv in self.items
            if g(inv, "ntsconfirmNum", "ntsConfirmNum") in nts_set
        ]
        out_dir = Path(self.path_var.get())
        if not messagebox.askyesno(
            "다운로드 확인",
            f"{len(selected_items)}건을 PDF로 저장합니다.\n저장 경로: {out_dir}\n\n진행하시겠습니까?",
        ):
            return

        self._set_busy(True)
        self.set_status(f"다운로드 준비 중... (0/{len(selected_items)})")
        threading.Thread(
            target=self._download_thread,
            args=(selected_items, out_dir),
            daemon=True,
        ).start()

    def _download_thread(self, selected_items: list[dict], out_dir: Path) -> None:
        def progress(n: int, total: int, nts: str) -> None:
            self.after(0, self.set_status, f"다운로드 {n}/{total}: {nts}")

        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            ok, fail = loop.run_until_complete(
                download_selected_items(self.svc, selected_items, progress, out_dir)
            )
            loop.close()
            self.after(0, self._on_download_done, ok, fail, out_dir)
        except Exception as e:
            self.after(0, self._on_download_error, str(e))

    def _on_download_done(self, ok: int, fail: int, out_dir: Path) -> None:
        self._set_busy(False)
        msg = f"저장 완료: 성공 {ok}건, 실패 {fail}건\n경로: {out_dir}"
        self.set_status(msg)
        messagebox.showinfo("완료", msg)

    def _on_download_error(self, msg: str) -> None:
        self._set_busy(False)
        self.set_status(f"저장 오류: {msg}")
        messagebox.showerror("저장 실패", msg)


def _check_update_thread(app: "PopbillApp") -> None:
    """백그라운드에서 GitHub Releases 최신 버전 확인."""
    import urllib.request
    try:
        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        req = urllib.request.Request(url, headers={"Accept": "application/vnd.github.v3+json"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        tag = data.get("tag_name", "").lstrip("vV")
        if not tag:
            return
        # 간단 비교: 점(.) 분리 숫자
        def ver_tuple(v: str):
            return tuple(int(x) for x in v.split(".") if x.isdigit())
        if ver_tuple(tag) > ver_tuple(APP_VERSION):
            download_url = data.get("html_url", "")
            # exe 다운로드 링크가 있으면 사용
            for asset in data.get("assets", []):
                if asset.get("name", "").endswith(".exe"):
                    download_url = asset["browser_download_url"]
                    break
            app.after(0, _show_update_dialog, app, tag, download_url)
    except Exception:
        pass  # 네트워크 오류 등은 무시


def _show_update_dialog(app: tk.Tk, new_ver: str, url: str) -> None:
    msg = (f"새 버전이 있습니다!\n\n"
           f"현재: v{APP_VERSION}  →  최신: v{new_ver}\n\n"
           f"다운로드 페이지를 열까요?")
    if messagebox.askyesno("업데이트 확인", msg, parent=app):
        import webbrowser
        webbrowser.open(url)


def main() -> None:
    app = PopbillApp()
    # 앱 시작 후 업데이트 확인 (백그라운드)
    threading.Thread(target=_check_update_thread, args=(app,), daemon=True).start()
    app.mainloop()


if __name__ == "__main__":
    main()
